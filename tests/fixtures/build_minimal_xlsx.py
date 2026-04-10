"""
Writes tests/fixtures/minimal.xlsx — smallest valid workbook for build_dashboard.py / CI.

Run from repo root:
  python tests/fixtures/build_minimal_xlsx.py
"""
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
OUT = Path(__file__).resolve().parent / "minimal.xlsx"

PIPELINE_HEADERS = [
    "Borrower",
    "Loan Officer",
    "Total Loan Amount",
    "Fast Pass",
    "Lender",
    "Purpose",
    "Loan Type",
    "Contract Close Date",
    "Actual Close Date",
    "Funded Date",
    "Interest Rate",
    "Processor",
]

FUNDED_HEADERS = [
    "Borrower",
    "Loan Officer",
    "Total Loan Amount",
    "Fast Pass",
    "Lender",
    "Purpose",
    "Loan Type",
    "Funded Date",
    "Interest Rate",
    "Processor",
]


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_p = wb.create_sheet("Loan Pipeline", 0)
    ws_p.append(PIPELINE_HEADERS)
    ws_p.append(
        [
            "Test Borrower One",
            "Jane Officer",
            350000,
            "Yes",
            "UWM",
            "Purchase",
            "Conventional",
            "2026-06-01",
            None,
            None,
            6.5,
            "Pat Processor",
        ]
    )

    ws_26 = wb.create_sheet("Apex Funded 2026", 1)
    ws_26.append(FUNDED_HEADERS)
    ws_26.append(
        [
            "Test Borrower Two",
            "Jane Officer",
            400000,
            "Yes",
            "Freedom",
            "Refi",
            "FHA",
            "2026-04-01",
            5.875,
            "Pat Processor",
        ]
    )

    ws_25 = wb.create_sheet("Apex Funded 2025", 2)
    ws_25.append(FUNDED_HEADERS)
    ws_25.append(
        [
            "Legacy Borrower",
            "Bob Officer",
            250000,
            "No",
            "Plaza",
            "Purchase",
            "VA",
            "2025-11-01",
            6.0,
            "Alex Processor",
        ]
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
