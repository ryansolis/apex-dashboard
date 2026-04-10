"""
Reads the loan spreadsheet and rebuilds index.html from the template.
Runs inside GitHub Actions after download_spreadsheet.py.

Local (no SharePoint): python tests/fixtures/build_minimal_xlsx.py && python build_dashboard.py --spreadsheet tests/fixtures/minimal.xlsx
"""
import argparse
import datetime
import json
import os
import sys

import openpyxl

# ── Config (override via env for new book years / sheet renames) ─────────────

PRIMARY_FUNDED_YEAR = int(os.environ.get("DASHBOARD_PRIMARY_FUNDED_YEAR", "2026"))
SHEET_LOAN_PIPELINE = os.environ.get("DASHBOARD_SHEET_PIPELINE", "Loan Pipeline")
SHEET_FUNDED_PRIMARY = os.environ.get(
    "DASHBOARD_SHEET_FUNDED_PRIMARY", f"Apex Funded {PRIMARY_FUNDED_YEAR}"
)
SHEET_FUNDED_PRIOR = os.environ.get(
    "DASHBOARD_SHEET_FUNDED_PRIOR", f"Apex Funded {PRIMARY_FUNDED_YEAR - 1}"
)

# Excel / Lotus 1900 date origin used by openpyxl (epoch serial 1 = 1900-01-01)
_EXCEL_ORIGIN = datetime.datetime(1899, 12, 30)


# ── Helpers ──────────────────────────────────────────────────────────────────


def parse_date(v, log_bad=None):
    """Normalize spreadsheet cell to YYYY-MM-DD or None (matches front-end intent)."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.date):
        return v.isoformat()
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        # Excel serial (whole or fractional days)
        if 1 <= float(v) <= 1_000_000:
            try:
                dt = _EXCEL_ORIGIN + datetime.timedelta(days=float(v))
                return dt.strftime("%Y-%m-%d")
            except (OverflowError, ValueError):
                pass
        if log_bad is not None:
            log_bad.append(("number", repr(v)))
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        if len(s) >= 10 and s[4] == "-" and s[7] == "-":
            return s[:10]
        if "/" in s:
            try:
                p = s.split("/")
                if len(p) == 3:
                    return f"{p[2]}-{p[0].zfill(2)}-{p[1].zfill(2)}"
            except Exception:
                pass
        if log_bad is not None:
            log_bad.append(("string", s[:80]))
        return None
    if log_bad is not None:
        log_bad.append(("other", type(v).__name__))
    return None


def to_num(v):
    try:
        return float(v)
    except Exception:
        return None


def fp_str(v):
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    return str(v).strip() if v else "N/A"


# ── Sheet readers ─────────────────────────────────────────────────────────────


def read_funded(wb, sheet, label_for_log):
    if sheet not in wb.sheetnames:
        print(f"  Warning: sheet '{sheet}' not found, skipping ({label_for_log}).")
        return []
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    rows = []
    bad_dates = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        d = dict(zip(headers, row))
        amt = to_num(d.get("Total Loan Amount"))
        if not amt:
            continue
        fd = parse_date(d.get("Funded Date"), bad_dates)
        rows.append(
            {
                "borrower": str(d.get("Borrower", "") or "").strip(),
                "loan_officer": str(d.get("Loan Officer", "") or "").strip(),
                "amount": amt,
                "fast_pass": fp_str(d.get("Fast Pass")),
                "lender": str(d.get("Lender", "") or "").strip(),
                "purpose": str(d.get("Purpose", "") or "").strip(),
                "loan_type": str(d.get("Loan Type", "") or "").strip(),
                "funded_date": fd,
                "rate": to_num(d.get("Interest Rate")),
                "processor": str(
                    d.get("Processor", "") or d.get("Loan Processor", "") or ""
                ).strip(),
            }
        )
    if bad_dates and os.environ.get("BUILD_VERBOSE"):
        sample = bad_dates[:5]
        print(f"  Note [{label_for_log}]: {len(bad_dates)} Funded Date cell(s) needed parsing help; sample: {sample}")
    return rows


def read_pipeline(wb):
    if SHEET_LOAN_PIPELINE not in wb.sheetnames:
        print(f"  Warning: sheet '{SHEET_LOAN_PIPELINE}' not found.")
        return []
    ws = wb[SHEET_LOAN_PIPELINE]
    headers = [c.value for c in ws[1]]
    rows = []
    bad_dates = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        b = row[0]
        if not b or not isinstance(b, str) or len(b.strip()) < 3:
            continue
        d = dict(zip(headers, row))
        amt = to_num(d.get("Total Loan Amount"))
        if not amt:
            continue
        rows.append(
            {
                "borrower": str(b).strip(),
                "loan_officer": str(d.get("Loan Officer", "") or "").strip(),
                "amount": amt,
                "fast_pass": fp_str(d.get("Fast Pass")),
                "lender": str(d.get("Lender", "") or "").strip(),
                "purpose": str(d.get("Purpose", "") or "").strip(),
                "loan_type": str(d.get("Loan Type", "") or "").strip(),
                "contract_close": parse_date(d.get("Contract Close Date"), bad_dates),
                "actual_close": parse_date(d.get("Actual Close Date"), bad_dates),
                "funded_date": parse_date(d.get("Funded Date"), bad_dates),
                "rate": to_num(d.get("Interest Rate")),
                "processor": str(
                    d.get("Processor", "") or d.get("Loan Processor", "") or ""
                ).strip(),
            }
        )
    if bad_dates and os.environ.get("BUILD_VERBOSE"):
        sample = bad_dates[:5]
        print(f"  Note [pipeline]: {len(bad_dates)} date cell(s) unparsed; sample: {sample}")
    return rows


# ── Main ──────────────────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(description="Build index.html from spreadsheet + template.")
    parser.add_argument(
        "--spreadsheet",
        default=os.environ.get("SPREADSHEET_PATH", "spreadsheet.xlsx"),
        help="Path to source .xlsx (default: spreadsheet.xlsx)",
    )
    parser.add_argument("--template", default="dashboard_template.html")
    parser.add_argument("--output", default="index.html")
    args = parser.parse_args()

    if not os.path.exists(args.spreadsheet):
        print(f"ERROR: {args.spreadsheet} not found. Run download_spreadsheet.py or --spreadsheet path to a fixture.")
        sys.exit(1)

    if not os.path.exists(args.template):
        print(f"ERROR: {args.template} not found.")
        sys.exit(1)

    print(f"Opening spreadsheet: {args.spreadsheet} …")
    wb = openpyxl.load_workbook(args.spreadsheet, data_only=True)

    data = {
        "pipeline": read_pipeline(wb),
        "funded2026": read_funded(wb, SHEET_FUNDED_PRIMARY, "funded primary"),
        "funded2025": read_funded(wb, SHEET_FUNDED_PRIOR, "funded prior"),
        "primaryYear": PRIMARY_FUNDED_YEAR,
        "meta": {
            "sheetLoanPipeline": SHEET_LOAN_PIPELINE,
            "sheetFundedPrimary": SHEET_FUNDED_PRIMARY,
            "sheetFundedPrior": SHEET_FUNDED_PRIOR,
        },
        "refreshed": datetime.datetime.now().strftime("%B %d, %Y at %I:%M %p"),
    }

    print(f"  pipeline:   {len(data['pipeline'])} loans")
    print(f"  funded primary ({SHEET_FUNDED_PRIMARY}): {len(data['funded2026'])} loans")
    print(f"  funded prior ({SHEET_FUNDED_PRIOR}): {len(data['funded2025'])} loans")

    with open(args.template, "r", encoding="utf-8") as f:
        template = f.read()

    upload_user = os.environ.get("DASHBOARD_UPLOAD_USER", "").strip()
    upload_pass = os.environ.get("DASHBOARD_UPLOAD_PASSWORD", "").strip()
    if "%%UPLOAD_GATE%%" in template:
        if upload_user and upload_pass:
            gate_js = "window.__UPLOAD_GATE__ = " + json.dumps(
                {"user": upload_user, "password": upload_pass}
            ) + ";\n"
        else:
            gate_js = ""
        template = template.replace("%%UPLOAD_GATE%%", gate_js)

    if "%%LOAN_DATA%%" not in template:
        print("ERROR: Template is missing the %%LOAN_DATA%% placeholder.")
        sys.exit(1)

    data_js = f"const RAW = {json.dumps(data, separators=(',', ':'))};"
    html = template.replace("%%LOAN_DATA%%", data_js)

    html = html.replace(
        "APEX<span>.</span>Mortgage",
        f'APEX<span>.</span>Mortgage <span style="font-size:11px;font-weight:400;color:var(--muted);margin-left:8px">Updated {data["refreshed"]}</span>',
        1,
    )

    with open(args.output, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"Dashboard written to {args.output} ({len(html):,} bytes)")
    print("Done.")


if __name__ == "__main__":
    main()
